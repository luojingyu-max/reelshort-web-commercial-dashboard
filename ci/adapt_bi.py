# -*- coding: utf-8 -*-
"""把新版 BI 4-表导出适配成引擎认的旧格式数据文件,并与历史增量合并。
用法: BI_FILE=/path/to/今日导出.xlsx  python3 ci/adapt_bi.py   (从仓库根目录运行)
      不设 BI_FILE 时,取当前目录下最新的 SEO数据看板*.xlsx / bi_*.xlsx。
合并规则: 每张表按"新文件该表的最早日期"为界 —— 保留现有 < 该日期的历史,追加新文件全部行。
新表→旧格式列位(已用 8/05 逐列验证):
 官网监控明细数据: 维度 日期/付费/国家/ALL; 老充值uv=新总付费uv(13),触达付费集uv=7,金币15,订阅17,总收入23,ltv 33-63
 交叉表1(策略):   日期/付费/国家/策略; 充值率插在8→后面右移1
 交叉表3(SKU):     日期/SKU/价格/类型/国家(16列原生)
 SEO监控明细(引流App): 维度2=官网引流APP; 支付成功uv11,订阅uv16,总收入25,广告31
"""
import openpyxl, warnings, re, os, glob
warnings.simplefilter("ignore")
def num(x):
    try: return float(x)
    except: return 0.0
D="ci/data"
NEW=os.environ.get("BI_FILE") or sorted(glob.glob("SEO数据看板*.xlsx")+glob.glob("bi_*.xlsx"))[-1]
print("适配输入:", NEW)
nb=openpyxl.load_workbook(NEW, read_only=True, data_only=True)
def is2026(v): return bool(re.match(r'2026-\d\d-\d\d', str(v)[:10]))
def newmin(sheet, r0):
    ds=[str(r[0])[:10] for r in sh(sheet).iter_rows(min_row=r0, values_only=True) if r and is2026(r[0])]
    return min(ds) if ds else "9999-99-99"

# BI 有时会改 sheet 名(交叉表1→官网商业化监控明细、交叉表3→档位支付明细),按别名+位置兜底
_ALIAS={"官网监控明细数据":["官网监控明细数据","官网监控明细"],
        "交叉表1":["交叉表1","官网商业化监控明细","策略交叉表"],
        "SEO监控明细数据":["SEO监控明细数据","SEO监控明细"],
        "交叉表3":["交叉表3","档位支付明细","SKU交叉表"]}
_POS={"官网监控明细数据":0,"交叉表1":1,"SEO监控明细数据":2,"交叉表3":3}
def sh(name):
    for a in _ALIAS.get(name,[name]):
        if a in nb.sheetnames: return nb[a]
    i=_POS.get(name)
    if i is not None and i < len(nb.sheetnames):
        print("  [warn] 未找到表名 %s,按位置取第%d个: %s"%(name,i+1,nb.sheetnames[i]))
        return nb.worksheets[i]
    raise KeyError(name)

# ---------- 1) 官网监控明细_recent.xlsx (2 行表头) ----------
# SKIP_MON=1 跳过监控明细(导出格式异常时,只更新其余3表,不污染大盘/国家)
if os.environ.get("SKIP_MON"):
    print("官网监控明细: 已跳过(SKIP_MON=1),大盘/国家保持原样")
else:
  MON=os.environ.get("MON_FILE")   # 监控明细可单独一个文件(否则用主工作簿的同名表)
  mon_ws=(openpyxl.load_workbook(MON, read_only=True, data_only=True).worksheets[0]) if MON else sh("官网监控明细数据")
  mall=list(mon_ws.iter_rows(min_row=1, values_only=True))
  mhdr=mall[1]                                     # 第2行=字段名
  mrows=[r for r in mall[2:] if r and is2026(r[0])]
  # 剔除残缺日:BI 当日数据未跑完时 DAU 会全为 0(有收入但无DAU),并入会污染大盘
  _dcol=next((i for i,v in enumerate(mhdr) if str(v).strip()=="DAU"), 4)
  _dau_by_day={}
  for _r in mrows: _dau_by_day[str(_r[0])[:10]]=_dau_by_day.get(str(_r[0])[:10],0)+num(_r[_dcol])
  _bad={d for d,v in _dau_by_day.items() if v<=0}
  if _bad:
      print("  [warn] 剔除残缺日(DAU=0,数据未跑完):", sorted(_bad))
      mrows=[r for r in mrows if str(r[0])[:10] not in _bad]
  cut=min((str(r[0])[:10] for r in mrows), default="9999-99-99")
  # 按表头名定位列(BI 常在中间插列,如"支付失败uv/率",按固定列号会全错)
  hidx={}
  for i,v in enumerate(mhdr):
      n=str(v).strip()
      if n and n not in hidx: hidx[n]=i
  def col(*names):
      for n in names:
          if n in hidx: return hidx[n]
      raise KeyError("监控明细缺列: "+"/".join(names))
  dau_i=col("DAU")
  _PAID={"未付费用户","已付费用户"}; _D0={"D0","非D0"}
  # 维度列 = DAU 之前的列;自动识别 付费/国家(D0 与 ALL 忽略,按 日期+付费+国家 聚合)
  dimc=list(range(1,dau_i))
  pcol=next((c for c in dimc if any(str(r[c]) in _PAID for r in mrows[:50])), 2)
  ccol=next((c for c in dimc if c!=pcol and not any(str(r[c]) in _D0 for r in mrows[:50])
             and len({str(r[c]) for r in mrows[:200]})>3), 1)
  SUM={"dau":col("DAU"),"view":col("观看uv"),"reach":col("触达付费集uv"),"order":col("创建订单uv"),
       "payok_uv":col("支付成功uv"),"payuv":col("总付费uv"),"coin":col("金币充值uv"),"sub":col("订阅uv"),
       "renew":col("续订uv"),"rev":col("总收入"),"subrev":col("订阅(续订)收入")}
  ltv0_i=col("ltv0")
  agg={}
  for r in mrows:
      k=(str(r[0])[:10], str(r[ccol]), str(r[pcol]))
      a=agg.get(k)
      if a is None: a=agg[k]={n:0.0 for n in SUM}; a["_ltv"]=[0.0]*31
      for n,ci in SUM.items(): a[n]+=num(r[ci])
      dau_r=num(r[dau_i])
      for j in range(31): a["_ltv"][j]+=num(r[ltv0_i+j])*dau_r
  sr=list(openpyxl.load_workbook(f"{D}/官网监控明细_recent.xlsx", read_only=True, data_only=True).worksheets[0].iter_rows(values_only=True))
  out=openpyxl.Workbook(); w=out.active; w.title="官网监控明细数据"
  for r in sr[:2]: w.append(list(r))
  kept=0
  for r in sr[2:]:
      if r and is2026(r[0]) and str(r[0])[:10]<cut: w.append(list(r)); kept+=1
  add=0
  for (d,c,p),a in sorted(agg.items()):
      o=[0]*63
      o[0]=d; o[1]=c; o[2]=p; o[3]="ALL"
      o[4]=a["dau"]; o[5]=a["view"]; o[7]=a["reach"]; o[9]=a["order"]
      o[11]=a["payuv"]                                          # 老充值uv=总付费uv
      o[12]=(a["payok_uv"]/a["order"]) if a["order"] else 0      # 支付成功率(重算)
      o[14]=a["coin"]; o[16]=a["sub"]; o[20]=a["renew"]
      o[22]=a["rev"]; o[25]=a["subrev"]
      o[30]=(a["rev"]/a["payuv"]) if a["payuv"] else 0           # ARPPU(重算)
      for j in range(31): o[32+j]=(a["_ltv"][j]/a["dau"]) if a["dau"] else 0   # LTV: DAU加权
      w.append(o); add+=1
  out.save(f"{D}/官网监控明细_recent.xlsx")
  print("官网监控明细 (来源%s DAU列%d 付费列%d 国家列%d 界<%s): 保留%d + 聚合后新增%d(原始%d行)"%(
      "MON_FILE" if MON else "工作簿", dau_i, pcol, ccol, cut, kept, add, len(mrows)))

# ---------- 2) 策略交叉表.xlsx (1 行表头) ----------
cut=newmin("交叉表1",2)
sr=list(openpyxl.load_workbook(f"{D}/策略交叉表.xlsx", read_only=True, data_only=True).worksheets[0].iter_rows(values_only=True))
out=openpyxl.Workbook(); w=out.active; w.title="策略交叉表"; w.append(list(sr[0])); kept=0
for r in sr[1:]:
    if r and is2026(r[0]) and str(r[0])[:10]<cut: w.append(list(r)); kept+=1
add=0
_x1=[r for r in sh("交叉表1").iter_rows(min_row=2, values_only=True) if r and is2026(r[0])]
# 自动识别 付费状态 在维度2还是维度3(BI 两种导出顺序都出现过)
_PAID2={"未付费用户","已付费用户"}
_pc=1 if any(str(r[1]) in _PAID2 for r in _x1[:50]) else 2
_cc=2 if _pc==1 else 1          # 国家列
print("  [交叉表1] 付费列=维度%d 国家列=维度%d 策略列=维度4"%(_pc+1,_cc+1))
for r in _x1:
    o=[0]*16
    o[0]=str(r[0])[:10]; o[1]=r[_pc]; o[2]=r[3]; o[3]=r[_cc]
    o[4]=r[4]; o[5]=r[5]; o[6]=r[6]; o[7]=r[7]; o[8]=r[9]; o[9]=r[10]
    o[10]=r[11]; o[11]=r[12]; o[12]=r[13]; o[13]=r[14]; o[14]=r[15]; o[15]=r[16]
    w.append(o); add+=1
out.save(f"{D}/策略交叉表.xlsx"); print("策略交叉表 (界<%s): 保留%d + 新增%d"%(cut,kept,add))

# ---------- 3) SKU交叉表.xlsx (统一16列含国家=col4; 历史行国家='全部') ----------
cut=newmin("交叉表3",2)
sr=list(openpyxl.load_workbook(f"{D}/SKU交叉表.xlsx", read_only=True, data_only=True).worksheets[0].iter_rows(values_only=True))
out=openpyxl.Workbook(); w=out.active; w.title="SKU交叉表"
hdr=list(sr[0]);
if len(hdr)==15: hdr=hdr[:4]+["国家"]+hdr[4:]
w.append(hdr); kept=0
for r in sr[1:]:
    if r and is2026(r[0]) and str(r[0])[:10]<cut:
        rr=list(r)
        if len(rr)==15: rr=rr[:4]+["全部"]+rr[4:]
        w.append(rr); kept+=1
_SKC={"美国","加拿大","澳大利亚","英国","法国","日本","意大利","巴西","墨西哥","智利","阿根廷","韩国","泰国"}  # SKU只看13策略国
add=0
# 按表头名取指标列(BI 曾在指标前插入"充值pv",固定列号会整体错位)
_x3=list(sh("交叉表3").iter_rows(min_row=1, values_only=True))
_h3={str(v).strip():i for i,v in enumerate(_x3[0]) if v}
_M3=["充值uv","金币充值uv","订阅uv","首订uv","续订uv","充值收入","金币充值收入","订阅(续订)收入","首订收入","续订收入","总收入"]
_idx3=[_h3[n] for n in _M3 if n in _h3]
if len(_idx3)!=11: raise SystemExit("交叉表3 缺列,仅匹配到%d/11: %s"%(len(_idx3),[n for n in _M3 if n not in _h3]))
print("  [交叉表3] 指标列位:",_idx3)
for r in _x3[1:]:
    if not r or not is2026(r[0]) or str(r[4]) not in _SKC: continue
    w.append([str(r[0])[:10],r[1],r[2],r[3],r[4]]+[round(num(r[i]),2) for i in _idx3]); add+=1
out.save(f"{D}/SKU交叉表.xlsx"); print("SKU交叉表 (界<%s): 保留%d + 新增%d(仅13策略国)"%(cut,kept,add))

# ---------- 4) 收入明细.xlsx (引流App=SEO监控明细口径; 合并历史) ----------
cut=newmin("SEO监控明细数据",3)
cols=["维度1","维度2","维度3","充值uv","金币充值uv","订阅uv","首订uv","续订uv","金币充值收入","订阅(续订)收入","首订收入","续订收入","总收入","广告收入"]
out=openpyxl.Workbook(); w=out.active; w.title="收入明细"; w.append(cols); kept=0
if os.path.exists(f"{D}/收入明细.xlsx"):
    for r in list(openpyxl.load_workbook(f"{D}/收入明细.xlsx", read_only=True, data_only=True).worksheets[0].iter_rows(values_only=True))[1:]:
        if r and is2026(r[0]) and str(r[0])[:10]<cut: w.append(list(r)); kept+=1
add=0
for r in sh("SEO监控明细数据").iter_rows(min_row=3, values_only=True):
    if not r or not is2026(r[0]): continue
    row=[0]*14
    row[0]=str(r[0])[:10]; row[1]="官网引流APP"; row[2]=r[2]
    row[3]=num(r[11]); row[5]=num(r[16]); row[12]=num(r[25]); row[13]=num(r[31])
    w.append(row); add+=1
out.save(f"{D}/收入明细.xlsx"); print("收入明细·引流App (界<%s): 保留%d + 新增%d"%(cut,kept,add))
