# -*- coding: utf-8 -*-
"""多源构建 payload2.json:
  data/官网大盘.xlsx            种子·官网整体日 (1/1-7/29)  —— site 历史
  data/官网监控明细_recent.xlsx  1531·国家×付费状态 (7/15-8/05) —— site 7/30起延伸 + 国家 + 二期后窗
  data/国家+付费.xlsx           种子·国家×付费状态 (~5/1-7/29,13国) —— 二期前窗 + 国家历史
  data/收入明细.xlsx            渠道日收入 (1/1-8/07,官网/引流) —— 官网vs引流
  strategy.json                一期/策略定义(不变)
口径:收入=金币充值+订阅(续订)+广告。列(4维表):DAU=4,充值uv=11,订阅uv=16,总收入=22,ltv0..30=32..62。"""
import json, datetime, re, warnings
warnings.simplefilter("ignore")
import openpyxl
from collections import defaultdict
D="data"
def ws(p): return openpyxl.load_workbook(p, read_only=True, data_only=True).worksheets[0]
def s2d(s):
    if s is None: return None
    if isinstance(s, datetime.datetime): return s.date().isoformat()
    if isinstance(s, datetime.date): return s.isoformat()
    ss=str(s).strip()
    if re.match(r'\d{4}-\d\d-\d\d',ss): return ss[:10]
    try: return (datetime.date(1899,12,30)+datetime.timedelta(days=int(float(ss)))).isoformat()
    except: return None
def num(x):
    try: return float(x)
    except: return 0.0

# ---------- 国家×付费状态 recs:seed(<7/15) + 1531(>=7/15) ----------
def load_cp(path):
    out=[]
    for r in ws(path).iter_rows(min_row=3, values_only=True):
        d=s2d(r[0])
        if not d or not r[1]: continue
        ltv=[num(r[32+k]) if len(r)>32+k else 0 for k in range(31)]
        out.append({"d":d,"c":r[1],"paid":(r[2]=="已付费用户"),"dau":num(r[4]),"view":num(r[5]),"reach":num(r[7]),"order":num(r[9]),
                    "uv":num(r[11]),"payok":num(r[12]),"coin":num(r[14]),"sub":num(r[16]),"renew":num(r[20]),"subrev":num(r[25]),"rev":num(r[22]),"ltv":ltv})
    return out
seed_cp=[x for x in load_cp(f"{D}/国家+付费.xlsx") if x["d"]<"2026-07-15"]
rec_cp=[x for x in load_cp(f"{D}/官网监控明细_recent.xlsx") if re.match(r'2026-\d\d-\d\d',x["d"])]
recs=seed_cp+rec_cp
maxd=max(x["d"] for x in recs); md=datetime.date.fromisoformat(maxd)

# ---------- site daily:seed官网大盘(<=7/29) + 1531聚合(>=7/30) ----------
site={}
for r in ws(f"{D}/官网大盘.xlsx").iter_rows(min_row=3, values_only=True):
    d=s2d(r[0])
    if not d or d>="2026-07-30": continue
    ltv=[num(r[32+k]) if len(r)>32+k else 0 for k in range(31)]
    site[d]={"dau":num(r[4]),"view":num(r[5]),"reach":num(r[7]),"order":num(r[9]),"uv":num(r[11]),"payok":num(r[12]),
             "coin":num(r[14]),"sub":num(r[16]),"subrev":num(r[25]),"rev":num(r[22]),"arppu":num(r[30]),"ltv":ltv}
agg=defaultdict(lambda:{"dau":0.0,"view":0.0,"reach":0.0,"order":0.0,"uv":0.0,"coin":0.0,"sub":0.0,"subrev":0.0,"rev":0.0,"pw":0.0,"lw":[0.0]*31})
for x in rec_cp:
    if x["d"]<"2026-07-30": continue
    a=agg[x["d"]]
    for k2 in ("dau","view","reach","order","uv","coin","sub","subrev","rev"): a[k2]+=x[k2]
    a["pw"]+=x["payok"]*x["order"]
    for k in range(31): a["lw"][k]+=x["ltv"][k]*x["dau"]
for d,a in agg.items():
    w=a["dau"] or 1
    site[d]={"dau":a["dau"],"view":a["view"],"reach":a["reach"],"order":a["order"],"uv":a["uv"],"coin":a["coin"],"sub":a["sub"],"subrev":a["subrev"],"rev":a["rev"],
             "payok":a["pw"]/a["order"] if a["order"] else 0,"arppu":a["rev"]/a["uv"] if a["uv"] else 0,"ltv":[a["lw"][k]/w for k in range(31)]}
dates=sorted(site)
dau=[round(site[d]["dau"]) for d in dates]
rev=[round(site[d]["rev"],2) for d in dates]
chargeuv=[round(site[d]["uv"]) for d in dates]
subuv=[round(site[d]["sub"]) for d in dates]
subrev_d=[round(site[d].get("subrev",0),2) for d in dates]
payrate=[round(site[d]["uv"]/site[d]["dau"]*100,4) if site[d]["dau"] else None for d in dates]
subrate=[round(site[d]["sub"]/site[d]["dau"]*100,4) if site[d]["dau"] else None for d in dates]
arppu=[round(site[d]["arppu"],4) for d in dates]
# mature LTV cohort (<= maxd-30)
cut=(md-datetime.timedelta(days=30)).isoformat(); mi=0
for i,d in enumerate(dates):
    if d<=cut: mi=i
ltv=[round(site[dates[mi]]["ltv"][k],4) for k in range(31)]; ltv_date=dates[mi]
def kpi(a):
    cur=a[-1]; prev=a[-2] if len(a)>1 else None
    return {"cur":cur,"pct":((cur-prev)/prev*100) if (cur is not None and prev) else None}
rev30=sum(v for v in rev[-30:] if v)

# ---------- 官网 vs 引流(收入明细·按国家聚合;直接取总收入列,缺失回退金币+订阅+广告) ----------
inc_site={}; inc_app={}
_rraw=defaultdict(lambda:{"s":0.0,"a":0.0,"sr":0.0,"ar":0.0,"apu":0.0,"asu":0.0,"aad":0.0})
for r in ws(f"{D}/收入明细.xlsx").iter_rows(min_row=2, values_only=True):
    d=s2d(r[0])
    if not d: continue
    r12=num(r[12]); t=r12 if r12>0 else num(r[8])+num(r[9])+num(r[13])
    if r[1]=="官网": _rraw[d]["s"]+=t; _rraw[d]["sr"]+=r12
    elif r[1]=="官网引流APP":
        _rraw[d]["a"]+=t; _rraw[d]["ar"]+=r12
        _rraw[d]["apu"]+=num(r[3]); _rraw[d]["asu"]+=num(r[5]); _rraw[d]["aad"]+=num(r[13])
for d,v in _rraw.items():
    if v["sr"]<=0 and v["ar"]<=0: continue   # 当天总收入列整列未回填(如最新一天,BI 也显示"—"),跳过防止掉成低值
    inc_site[d]=round(v["s"],2); inc_app[d]=round(v["a"],2)
# 官网线改从「官网数据」(site daily·监控明细/种子)取 → 收入数据只需含引流app
ov_dates=[d for d in sorted(inc_app) if d>="2026-06-01"]
ov_site=[round(site[d]["rev"],2) if d in site else None for d in ov_dates]
ov_app=[round(inc_app.get(d,0),2) for d in ov_dates]
app=[{"date":d,"rev":round(inc_app.get(d,0),2),"charge":round(inc_app.get(d,0)-_rraw[d]["aad"],2),"ad":round(_rraw[d]["aad"],2),
      "pay_uv":round(_rraw[d]["apu"]),"sub_uv":round(_rraw[d]["asu"])} for d in ov_dates]

# ---------- targets / dash_mom (site) ----------
targets={"2026-07":6*10000,"2026-08":8*10000,"2026-09":10*10000}
cur_month=dates[-1][:7]; mtd=sum(v for dd,v in zip(dates,rev) if dd[:7]==cur_month and v); target_cur=targets.get(cur_month)
def sw(a,b):
    ix=[i for i,dd in enumerate(dates) if a<=dd<=b]
    R=sum(rev[i] or 0 for i in ix); DA=sum(dau[i] or 0 for i in ix); UV=sum(chargeuv[i] or 0 for i in ix); n=len(ix) or 1
    return {"rev":R,"iap_day":R/n,"arpu":R/DA if DA else 0,"arppu":R/UV if UV else 0,"payrate":UV/DA*100 if DA else 0}
iso=lambda x:x.isoformat()
tmv=(iso(md.replace(day=1)),dates[-1]); lme=md.replace(day=1)-datetime.timedelta(days=1)
lmv=(iso(lme.replace(day=1)),iso(lme.replace(day=min(md.day,lme.day))))
wkv=(iso(md-datetime.timedelta(days=6)),dates[-1]); pwv=(iso(md-datetime.timedelta(days=13)),iso(md-datetime.timedelta(days=7)))
TM,LM,Wk,PW=sw(*tmv),sw(*lmv),sw(*wkv),sw(*pwv)
pct=lambda c,p: round((c-p)/p*100,1) if p else None
dash_metrics=[{"key":k,"label":l,"fmt":f,"cur":round(TM[k],4),"mom":pct(TM[k],LM[k]),"wow":pct(Wk[k],PW[k])}
 for k,l,f in [("rev","收入(本月至今)","usd"),("arpu","ARPU","usd4"),("arppu","ARPPU","usd1"),("payrate","付费率","pct3"),("iap_day","日均IAP$","usd")]]
dash_mom={"metrics":dash_metrics,"win":{"month":list(tmv),"lastmonth":list(lmv),"week":list(wkv),"pastweek":list(pwv)}}

# ---------- 国家 ctab / LTV / detail ----------
def win(a,b): return [x for x in recs if a<=x["d"]<=b]
def arev(rows,c,paid=None): return sum(x["rev"] for x in rows if x["c"]==c and (paid is None or x["paid"]==paid))
def cagg(rows,c,paid=None):
    d_=u_=s_=r_=v_=rc_=o_=sr_=rn_=0.0; days=set()
    for x in rows:
        if x["c"]!=c or (paid is not None and x["paid"]!=paid): continue
        d_+=x["dau"]; u_+=x["uv"]; s_+=x["sub"]; r_+=x["rev"]; v_+=x["view"]; rc_+=x["reach"]; o_+=x["order"]; sr_+=x["subrev"]; rn_+=x["renew"]; days.add(x["d"])
    nd=len(days) or 1
    return {"rev":round(r_,2),"subrev":round(sr_,2),"dau":round(d_/nd),"payrate":round(u_/d_*100,3) if d_ else 0,
            "subrate":round(s_/d_*100,3) if d_ else 0,"renewrate":round(rn_/s_*100,2) if s_ else 0,"arppu":round(r_/u_,2) if u_ else 0,
            "viewrate":round(v_/d_*100,2) if d_ else 0,"reachrate":round(rc_/v_*100,2) if v_ else 0,
            "orderrate":round(o_/rc_*100,2) if rc_ else 0,"arpu":round(r_/d_,4) if d_ else 0}
tm=(iso(md.replace(day=1)),maxd); lm=(iso(lme.replace(day=1)),iso(lme.replace(day=min(md.day,lme.day))))
wk=(iso(md-datetime.timedelta(days=6)),maxd); pw=(iso(md-datetime.timedelta(days=13)),iso(md-datetime.timedelta(days=7)))
TMc,LMc,Wc,PWc=win(*tm),win(*lm),win(*wk),win(*pw)
# 日环比: 最新日 vs 其前一个有数据日(数据可能跳天,故取实际存在的前一日)
_cds=sorted({x["d"] for x in recs}); dcur=_cds[-1]; dprev=_cds[-2] if len(_cds)>1 else None
Dc=win(dcur,dcur); DPc=win(dprev,dprev) if dprev else []
countries=sorted({x["c"] for x in recs}, key=lambda c:-arev(recs,c))
ctab=[]
for c in countries:
    cur=cagg(TMc,c); paid=cagg(TMc,c,True); unpaid=cagg(TMc,c,False)
    ctab.append({"c":c,**cur,"rev_paid":paid["rev"],"rev_unpaid":unpaid["rev"],
                 "rev_mom":pct(arev(TMc,c),arev(LMc,c)),"rev_wow":pct(arev(Wc,c),arev(PWc,c)),
                 "rev_dod":pct(arev(Dc,c),arev(DPc,c)) if dprev else None})
cwindows={"month":list(tm),"lastmonth":list(lm),"week":list(wk),"pastweek":list(pw),
          "day":[dcur,dcur],"prevday":([dprev,dprev] if dprev else None)}
mcut=(md-datetime.timedelta(days=30)).isoformat()
def cltv(c):
    rs=[x for x in recs if x["c"]==c and x["d"]<=mcut]; w=sum(x["dau"] for x in rs)
    if not w: return [None]*31
    return [round(sum(x["ltv"][k]*x["dau"] for x in rs)/w,4) for k in range(31)]
country_ltv={"countries":countries,"curve":{c:cltv(c) for c in countries},"mature_cut":mcut}
country_ltv_table=[{"c":c,"ltv0":country_ltv["curve"][c][0],"ltv7":country_ltv["curve"][c][7],
                    "ltv14":country_ltv["curve"][c][14],"ltv30":country_ltv["curve"][c][30]} for c in countries]
# ---------- 各国漏斗率日趋势(收入 Top11) ----------
fstart=(md-datetime.timedelta(days=44)).isoformat()
funnel_dates=sorted({x["d"] for x in recs if x["d"]>=fstart})
_grp=[("美加澳英",["美国","加拿大","澳大利亚","英国"]),("法/意/日",["法国","意大利","日本"]),
      ("墨/智/阿",["墨西哥","智利","阿根廷"]),("巴西",["巴西"]),("韩/泰",["韩国","泰国"])]
_cset={x["c"] for x in recs}
funnel_cty=[c for _,cs in _grp for c in cs if c in _cset]
_fa=defaultdict(lambda:{"v":0.0,"rc":0.0,"o":0.0,"p":0.0,"d":0.0})
for x in recs:
    if x["c"] in funnel_cty and x["d"]>=fstart:
        a=_fa[(x["c"],x["d"])]; a["v"]+=x["view"]; a["rc"]+=x["reach"]; a["o"]+=x["order"]; a["p"]+=x["uv"]; a["d"]+=x["dau"]
def _raw(c,m): return [round(_fa[(c,d)][m]) if (c,d) in _fa else 0 for d in funnel_dates]
funnel={"dates":funnel_dates,
        "groups":[{"name":n,"countries":[c for c in cs if c in _cset]} for n,cs in _grp],
        "raw":{c:{"view":_raw(c,"v"),"reach":_raw(c,"rc"),"order":_raw(c,"o"),"pay":_raw(c,"p"),"dau":_raw(c,"d")} for c in funnel_cty}}
detail=[]
for x in sorted(recs,key=lambda r:(r["d"],r["c"])):
    detail.append([x["d"],x["c"],"已付费" if x["paid"] else "未付费",round(x["dau"]),round(x["uv"]),round(x["sub"]),
                   round(x["rev"],2),round(x["ltv"][0],4),round(x["ltv"][7],4),round(x["ltv"][14],4),round(x["ltv"][30],4)])
detail_cols=["日期","国家","付费状态","DAU","充值uv","订阅uv","总收入","LTV0","LTV7","LTV14","LTV30"]

# ---------- 二期(前 seed 6.17-7.16 / 后 1531 7.17-8.05) ----------
PRE=("2026-06-17","2026-07-16"); POST=("2026-07-17",maxd)
p2c=["美国","澳大利亚","意大利","墨西哥","英国","巴西","日本","法国","加拿大","智利","阿根廷"]
def seg(src,c,paid,w):
    rs=[x for x in src if x["c"]==c and x["paid"]==paid and w[0]<=x["d"]<=w[1]]
    days=len(set(x["d"] for x in rs)) or 1; wt=sum(x["dau"] for x in rs) or 1
    da=sum(x["dau"] for x in rs)/days; uv=sum(x["uv"] for x in rs)/days; sb=sum(x["sub"] for x in rs)/days; rv=sum(x["rev"] for x in rs)/days
    lk=lambda k: sum(x["ltv"][k]*x["dau"] for x in rs)/wt
    return da,uv,sb,rv,lk
def build2(paid):
    out=[]
    for c in p2c:
        da,ua,sa,ra,la=seg(seed_cp,c,paid,PRE); db,ub,sb,rb,lb=seg(rec_cp,c,paid,POST)
        pr=lambda u,d: round(u/d*100,3) if d else 0
        out.append({"c":c,"pr_a":pr(ua,da),"pr_b":pr(ub,db),"sr_a":pr(sa,da),"sr_b":pr(sb,db),
            "arppu_a":round(ra/ua,1) if ua else 0,"arppu_b":round(rb/ub,1) if ub else 0,
            "iap_a":round(ra),"iap_b":round(rb),"iapchg":("+" if rb>=ra else "")+(f"{(rb-ra)/ra*100:.0f}%" if ra else "N/A"),
            "l0_a":round(la(0),4),"l0_b":round(lb(0),4),"l7_a":round(la(7),4),"l7_b":round(lb(7),4),
            "l14_a":round(la(14),4),"l14_b":round(lb(14),4),"l30_a":round(la(30),4),"l30_b":round(lb(30),4)})
    return out
phase2_unpaid=build2(False); phase2_paid=build2(True)
pts=[0,1,7,14,30]
def segagg(src,paid,w):
    rs=[x for x in src if x["paid"]==paid and x["c"] in p2c and w[0]<=x["d"]<=w[1]]; wt=sum(x["dau"] for x in rs) or 1
    return [round(sum(x["ltv"][k]*x["dau"] for x in rs)/wt,4) for k in pts]
phase2_ltv={"pts":pts,"unpaid_pre":segagg(seed_cp,False,PRE),"unpaid_post":segagg(rec_cp,False,POST),
            "paid_pre":segagg(seed_cp,True,PRE),"paid_post":segagg(rec_cp,True,POST)}

# ---------- site_detail (大盘明细) ----------
sd_cols=["日期","DAU","观看uv","观看率%","触达付费集uv","付费集触达率%","创建订单uv","创建订单率%","充值uv","充值成功率%","充值率%","金币充值uv","金币充值占比%",
         "订阅uv","订阅uv占比%","订阅率%","总收入","总收入ARPU","总收入ARPPU","LTV0","LTV7","LTV14","LTV30"]
site_detail=[]
for d in dates:
    s=site[d]; da=s["dau"] or 1; uv=s["uv"] or 1; vw=s["view"] or 1; rh=s.get("reach",0) or 1
    site_detail.append([d, round(s["dau"]), round(s["view"]), round(s["view"]/da*100,3),
        round(s.get("reach",0)), round(s.get("reach",0)/vw*100,2), round(s["order"]), round(s["order"]/rh*100,2),
        round(s["uv"]), round(s["payok"]*100,2), round(s["uv"]/da*100,4), round(s["coin"]), round(s["coin"]/uv*100,2),
        round(s["sub"]), round(s["sub"]/uv*100,2), round(s["sub"]/da*100,4), round(s["rev"],2), round(s["rev"]/da,4), round(s["arppu"],2),
        round(s["ltv"][0],4), round(s["ltv"][7],4), round(s["ltv"][14],4), round(s["ltv"][30],4)])

strat=json.load(open("strategy.json"))
# 一期加入韩国、泰国(6.18 上架·注册国家):文档表2只有11国,韩/泰按同结构从种子算(前6.03-6.16/后6.18-7.01)
def p1row(c):
    def sg(w):
        rs=[x for x in seed_cp if x["c"]==c and w[0]<=x["d"]<=w[1]]
        days=len(set(x["d"] for x in rs)) or 1; wt=sum(x["dau"] for x in rs) or 1
        da=sum(x["dau"] for x in rs)/days; uv=sum(x["uv"] for x in rs)/days; sb=sum(x["sub"] for x in rs)/days; rv=sum(x["rev"] for x in rs)/days
        return da,uv,sb,rv,(lambda k: sum(x["ltv"][k]*x["dau"] for x in rs)/wt)
    da,ua,sa,ra,La=sg(("2026-06-03","2026-06-16")); db,ub,sb2,rb,Lb=sg(("2026-06-18","2026-07-01"))
    pr=lambda u,d: round(u/d*100,3) if d else 0
    chg=lambda a,b: ("+" if b>=a else "")+(f"{(b-a)/a*100:.0f}%" if a else "N/A")
    return [c, round(da), round(db), pr(ua,da), pr(ub,db), pr(sa,da), pr(sb2,db),
            round(ua,1), round(ub,1), round(ra/ua,1) if ua else 0, round(rb/ub,1) if ub else 0,
            round(ra), round(rb), chg(ra,rb),
            round(La(0),3), round(Lb(0),3), round(La(7),3), round(Lb(7),3),
            round(La(14),3), round(Lb(14),3), round(La(30),3), round(Lb(30),3), chg(La(30),Lb(30))]
panel1=strat["panel1"]+[p1row("韩国"), p1row("泰国")]
# ---------- 面板策略明细(交叉表:日期×货架ID×策略) ----------
P2C13=["美国","加拿大","澳大利亚","英国","法国","日本","意大利","巴西","墨西哥","智利","阿根廷","韩国","泰国"]
sd_rows=[]; srev=defaultdict(lambda:{"p":0.0,"u":0.0}); sset=set()
cbys=defaultdict(lambda:defaultdict(lambda:{"exp":0.0,"pay":0.0,"fo":0.0,"rev":0.0}))
for r in ws(f"{D}/策略交叉表.xlsx").iter_rows(min_row=2, values_only=True):
    d=s2d(r[0])
    if not d or not re.match(r'2026-\d\d-\d\d',str(r[0])[:10]) or not r[2]: continue
    pd="已付费" if r[1]=="已付费用户" else "未付费"
    sd_rows.append([d, pd, r[2], r[3], round(num(r[4])), round(num(r[5])), round(num(r[6])), round(num(r[7])),
        round(num(r[8])), round(num(r[9])), round(num(r[10])), round(num(r[11])), round(num(r[12]),2),
        round(num(r[13]),2), round(num(r[14]),2), round(num(r[15]))])
    srev[r[2]]["p" if pd=="已付费" else "u"]+=num(r[12]); sset.add(r[2])
    if r[3] in P2C13:
        cbys[r[3]][r[2]]["exp"]+=num(r[5]); cbys[r[3]][r[2]]["pay"]+=num(r[7]); cbys[r[3]][r[2]]["fo"]+=num(r[11]); cbys[r[3]][r[2]]["rev"]+=num(r[12])
def _sl(s): return s.replace("官网-","").replace("-kim","").replace("kim ","")
strat_by_country={}; strat_bubble=[]
for c in P2C13:
    its=[{"sl":_sl(s),"exp":round(v["exp"]),"rate":round(v["pay"]/v["exp"]*100,2) if v["exp"] else 0,"subr":round(v["fo"]/v["pay"]*100,2) if v["pay"] else 0,"rev":round(v["rev"],2)}
         for s,v in cbys[c].items() if v["exp"]>0]
    its.sort(key=lambda x:-x["exp"])
    if its: strat_by_country[c]=its[:6]
    for it in its:
        if it["exp"]>=20: strat_bubble.append({"c":c,"sl":it["sl"],"exp":it["exp"],"rate":it["rate"],"rev":it["rev"]})
sd_rows.sort(key=lambda x:(x[0],x[2],x[1],x[3]))
strat_detail_cols=["日期","付费状态","策略","注册国家","曝光pv","曝光uv","充值pv","充值uv","金币充值pv","金币充值uv","首订pv","首订uv","总收入","金币充值收入","首订收入","付费后播放uv"]
strat_list=sorted(sset)
strat_rev=sorted([{"s":k,"paid":round(v["p"],2),"unpaid":round(v["u"],2),"total":v["p"]+v["u"]} for k,v in srev.items()], key=lambda x:-x["total"])[:12]

# ---------- SKU 趋势(SKU交叉表 6/1-8/9,日期×SKU×价格×商品类型)· 金币 vs 订阅 ----------
import os as _os
sku_dates=[]; sku_coin=[]; sku_sub=[]; sku_detail=[]; sku_list=[]; sku_countries=[]; sku_nodes=[]; sku_summary={}
_skf=f"{D}/SKU交叉表.xlsx"
if _os.path.exists(_skf):
    sk_c=defaultdict(lambda:defaultdict(float)); sk_s=defaultdict(lambda:defaultdict(float))
    tc=defaultdict(float); ts=defaultdict(float); _ds=set()
    def _win(d):
        if d<'2026-06-17': return 'w0'
        if d<='2026-07-16': return 'w1'
        if d<='2026-08-07': return 'w2'
        return 'w3'
    _wd={'w0':16,'w1':30,'w2':22,'w3':2}; _sm=defaultdict(lambda:defaultdict(float))
    for r in ws(_skf).iter_rows(min_row=2, values_only=True):
        d=s2d(r[0])
        if not d or not r[1] or not re.match(r'2026-\d\d-\d\d',d): continue
        nm=str(r[1]).strip(); t4=str(r[3]); ctry=str(r[4]) if r[4] else "全部"; _ds.add(d)
        coin=t4.startswith('0'); typ='金币' if coin else ('订阅续订' if t4.startswith('2') else '订阅首购')
        _rv=num(r[15])
        if coin: sk_c[nm][d]+=_rv; tc[nm]+=_rv
        else: sk_s[nm][d]+=_rv; ts[nm]+=_rv
        _sm['金币' if coin else '订阅'][_win(d)]+=_rv
        sku_detail.append([d,nm,typ,round(num(r[2])/100,2),ctry,round(num(r[5])),round(num(r[6])),round(num(r[7])),
            round(num(r[8])),round(num(r[9])),round(num(r[10]),2),round(num(r[11]),2),round(num(r[12]),2),
            round(num(r[13]),2),round(num(r[14]),2),round(num(r[15]),2)])
    sku_dates=sorted(_ds)
    _tc=[k for k,_ in sorted(tc.items(), key=lambda kv:-kv[1])[:6]]
    _ts=[k for k,_ in sorted(ts.items(), key=lambda kv:-kv[1])[:6]]
    sku_coin=[{"name":k,"data":[round(sk_c[k].get(d,0),2) for d in sku_dates]} for k in _tc]
    sku_sub=[{"name":k,"data":[round(sk_s[k].get(d,0),2) for d in sku_dates]} for k in _ts]
    sku_detail.sort(key=lambda x:(x[0], -x[15]))
    sku_list=sorted(set(list(tc)+list(ts))); sku_countries=sorted({str(x[4]) for x in sku_detail})
    sku_nodes=[{"date":"2026-06-17","label":"6.17 一期"},{"date":"2026-07-17","label":"7.17 二期"},{"date":"2026-08-08","label":"8.8 三期改金币档"}]
    sku_summary={"wins":["一期前","一期6.17","二期7.17","三期8.8"],
        "rows":[{"k":k,"vals":[round(_sm[k][w]/_wd[w]) for w in ['w0','w1','w2','w3']]} for k in ['金币','订阅']]}
sku_detail_cols=["日期","SKU","类型","价格$","国家","充值uv","金币充值uv","订阅uv","首订uv","续订uv","充值收入","金币充值收入","订阅收入","首订收入","续订收入","总收入"]

# ---------- 时间范围标注(各卡片用) ----------
def _rg(a,b): return (str(a)+" ~ "+str(b)) if (a and b) else ""
_sdd=sorted({r[0] for r in sd_rows}) if sd_rows else []
ranges={"dash":_rg(dates[0],dates[-1]),
        "ov":_rg(ov_dates[0],ov_dates[-1]) if ov_dates else "",
        "country":_rg(tm[0],tm[1]),"country_prev":_rg(lm[0],lm[1]),
        "cltv":"成熟批次 · 注册日 ≤ "+mcut,"ltv_date":ltv_date,
        "strat":_rg(_sdd[0],_sdd[-1]) if _sdd else "",
        "sku":_rg(sku_dates[0],sku_dates[-1]) if sku_dates else "",
        "phase2_pre":_rg(*PRE),"phase2_post":_rg(*POST),"maxd":maxd}

P={"gen":dates[-1],"dates":dates,"dau":dau,"rev":rev,"payrate":payrate,"subrate":subrate,"arppu":arppu,
   "chargeuv":chargeuv,"subuv":subuv,"subrev_d":subrev_d,"ltv":ltv,"ltv_date":ltv_date,
   "kpi":{"dau":kpi(dau),"payrate":kpi(payrate),"arppu":kpi(arppu),"rev":kpi(rev),"ltv30":ltv[30],"rev30":rev30},
   "app":app,"ov_dates":ov_dates,"ov_site":ov_site,"ov_app":ov_app,
   "targets":targets,"cur_month":cur_month,"mtd":round(mtd),"target_cur":target_cur,"dash_mom":dash_mom,
   "ctab":ctab,"cwindows":cwindows,"panel1":panel1,"panel1_header":strat["panel1_header"],
   "phase2_unpaid":phase2_unpaid,"phase2_paid":phase2_paid,"strategy":strat["strategy"],"strategy_header":strat["strategy_header"],
   "country_ltv":country_ltv,"country_ltv_table":country_ltv_table,"phase2_ltv":phase2_ltv,"funnel":funnel,
   "detail":detail,"detail_cols":detail_cols,"detail_countries":countries,
   "site_detail_cols":sd_cols,"site_detail":site_detail,
   "strat_detail":sd_rows,"strat_detail_cols":strat_detail_cols,"strat_list":strat_list,"strat_rev":strat_rev,
   "strat_by_country":strat_by_country,"strat_bubble":strat_bubble,
   "sku_dates":sku_dates,"sku_coin":sku_coin,"sku_sub":sku_sub,"sku_nodes":sku_nodes,"sku_summary":sku_summary,
   "sku_detail":sku_detail,"sku_detail_cols":sku_detail_cols,"sku_list":sku_list,"sku_countries":sku_countries,
   "ranges":ranges}
json.dump(P,open("payload2.json","w"),ensure_ascii=False)
print("built: site",dates[0],"->",dates[-1],"(",len(dates),"d) | latest DAU",dau[-1],"| 7月rev",round(sum(v for dd,v in zip(dates,rev) if dd[:7]=='2026-07')),
      "| curmonth",cur_month,"MTD",round(mtd),"| ctab",len(ctab),"| detail",len(detail),"| 国家数",len(countries))
